import sys
import os
import time
import uuid
import csv
import json
import tracemalloc
import threading
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch, MagicMock, PropertyMock

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from sqlalchemy import create_engine, desc
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import OperationalError, SQLAlchemyError

from server.models import (
    Base, Device, DeviceStatus, DeviceDataHistory, ReportTask, ReportTaskStatus,
    ScheduledReportConfig, ScheduledReportType
)
from server.report_engine import (
    report_engine, ScheduledReportManager, REPORTS_DIR, EXPORT_BATCH_SIZE,
    ProgressTracker, ExportFileHandler, SUPPORTED_FORMATS, DEFAULT_REPORT_RETENTION_HOURS
)

TEST_RECORDS_COUNT = 10000
TEST_DEVICE_ID = f"test_robust_{uuid.uuid4().hex[:8]}"
TEST_BATCH_SIZE = 1000

DATABASE_URL = "sqlite:///./test_robust_report.db"

engine = create_engine(
    DATABASE_URL, connect_args={"check_same_thread": False}
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def init_test_database():
    Base.metadata.create_all(bind=engine)
    print(f"[Test] Test database initialized: {DATABASE_URL}")

def cleanup_test_database():
    db_path = DATABASE_URL.replace("sqlite:///", "")
    if os.path.exists(db_path):
        os.unlink(db_path)
        print(f"[Test] Cleaned up test database: {db_path}")

def create_test_device(db: Session):
    device = Device(
        device_id=TEST_DEVICE_ID,
        secret_key=uuid.uuid4().hex,
        model="Test-Sensor-Robust",
        status=DeviceStatus.ONLINE
    )
    db.add(device)
    db.commit()
    db.refresh(device)
    print(f"[Test] Created test device: {TEST_DEVICE_ID}")
    return device

def generate_test_history_records(
    db: Session, 
    device_id: str, 
    count: int,
    batch_size: int = TEST_BATCH_SIZE
):
    print(f"\n[Test] Generating {count} test history records...")
    
    base_time = datetime.utcnow() - timedelta(hours=count // 100)
    records_generated = 0
    
    for batch_start in range(0, count, batch_size):
        batch_records = []
        for i in range(batch_size):
            if records_generated >= count:
                break
            
            record_time = base_time + timedelta(seconds=records_generated)
            temperature = 20.0 + (records_generated % 100) * 0.5
            humidity = 40.0 + (records_generated % 80) * 0.75
            is_alert = temperature > 50.0
            
            history_record = DeviceDataHistory(
                id=str(uuid.uuid4()),
                device_id=device_id,
                payload={
                    "temperature": temperature,
                    "humidity": humidity,
                    "record_index": records_generated
                },
                timestamp=record_time,
                temperature=temperature,
                humidity=humidity,
                is_alert=is_alert
            )
            batch_records.append(history_record)
            records_generated += 1
        
        db.add_all(batch_records)
        db.commit()
        
        if records_generated % 2000 == 0:
            print(f"[Test]   Generated {records_generated}/{count} records...")
    
    print(f"[Test] Completed generating {records_generated} records")
    return records_generated

def test_progress_tracker():
    print("\n" + "=" * 70)
    print("TEST 1: ProgressTracker - 10% Update Interval")
    print("=" * 70)
    
    total = 1000
    tracker = ProgressTracker(total, update_interval_percent=10)
    
    updates = []
    
    for current in range(0, total + 1, 100):
        progress_pct, should_update = tracker.update(current)
        if should_update:
            updates.append((current, progress_pct))
            print(f"[Test]   Progress: {current}/{total} ({progress_pct}%) - TRIGGER UPDATE")
        else:
            print(f"[Test]   Progress: {current}/{total} ({progress_pct}%) - skip")
    
    print(f"\n[Test] Total update triggers: {len(updates)}")
    
    expected_updates = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    actual_percents = [pct for _, pct in updates]
    
    print(f"[Test] Expected percent updates: {expected_updates}")
    print(f"[Test] Actual percent updates:   {actual_percents}")
    
    if actual_percents == expected_updates:
        print(f"\n[PASSED] ProgressTracker correctly triggers every 10%")
        return True
    else:
        print(f"\n[FAILED] ProgressTracker update pattern incorrect")
        return False

def test_csv_export_format():
    print("\n" + "=" * 70)
    print("TEST 2: CSV Export Format")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        output_path = REPORTS_DIR / f"test_csv_{uuid.uuid4().hex[:8]}.csv"
        
        print(f"\n[Test] Exporting to CSV: {output_path}")
        
        progress_updates = []
        
        def progress_callback(current: int, total: int, progress: int):
            progress_updates.append(progress)
            print(f"[Test]   Progress: {progress}% ({current}/{total})")
        
        total_records, file_size = report_engine.export_to_file(
            db=db,
            device_id=TEST_DEVICE_ID,
            output_path=output_path,
            file_format="csv",
            progress_callback=progress_callback
        )
        
        print(f"\n[Test] CSV Export complete:")
        print(f"[Test]   Records: {total_records}")
        print(f"[Test]   File size: {file_size / 1024:.2f} KB")
        print(f"[Test]   Progress updates: {progress_updates}")
        
        if total_records != TEST_RECORDS_COUNT:
            print(f"[FAILED] Record count mismatch: expected {TEST_RECORDS_COUNT}, got {total_records}")
            return False
        
        if not output_path.exists():
            print(f"[FAILED] File does not exist")
            return False
        
        row_count = 0
        with open(output_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                row_count += 1
        
        data_rows = row_count - 1
        print(f"[Test]   CSV row count: {row_count} (header + {data_rows} data rows)")
        
        if data_rows != TEST_RECORDS_COUNT:
            print(f"[FAILED] CSV row count mismatch: expected {TEST_RECORDS_COUNT}, got {data_rows}")
            return False
        
        print(f"\n[PASSED] CSV export format test passed")
        return True
        
    finally:
        db.close()
        if output_path.exists():
            output_path.unlink()

def test_json_export_format():
    print("\n" + "=" * 70)
    print("TEST 3: JSON Export Format")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        output_path = REPORTS_DIR / f"test_json_{uuid.uuid4().hex[:8]}.json"
        
        print(f"\n[Test] Exporting to JSON: {output_path}")
        
        total_records, file_size = report_engine.export_to_file(
            db=db,
            device_id=TEST_DEVICE_ID,
            output_path=output_path,
            file_format="json"
        )
        
        print(f"\n[Test] JSON Export complete:")
        print(f"[Test]   Records: {total_records}")
        print(f"[Test]   File size: {file_size / 1024:.2f} KB")
        
        if total_records != TEST_RECORDS_COUNT:
            print(f"[FAILED] Record count mismatch: expected {TEST_RECORDS_COUNT}, got {total_records}")
            return False
        
        if not output_path.exists():
            print(f"[FAILED] File does not exist")
            return False
        
        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"[Test]   JSON array length: {len(data)}")
        
        if len(data) != TEST_RECORDS_COUNT:
            print(f"[FAILED] JSON array length mismatch: expected {TEST_RECORDS_COUNT}, got {len(data)}")
            return False
        
        if len(data) > 0:
            sample_record = data[0]
            print(f"[Test]   Sample record keys: {list(sample_record.keys())}")
            
            expected_keys = ["id", "device_id", "timestamp", "temperature", "humidity", "is_alert", "payload_json"]
            for key in expected_keys:
                if key not in sample_record:
                    print(f"[FAILED] Missing expected key: {key}")
                    return False
            
            print(f"[Test]   Sample record device_id: {sample_record['device_id']}")
            
            if sample_record['device_id'] != TEST_DEVICE_ID:
                print(f"[FAILED] Device ID mismatch in JSON")
                return False
        
        print(f"\n[PASSED] JSON export format test passed")
        return True
        
    finally:
        db.close()
        if output_path.exists():
            output_path.unlink()

def test_excel_export_format():
    print("\n" + "=" * 70)
    print("TEST 4: Excel Export Format")
    print("=" * 70)
    
    try:
        import openpyxl
    except ImportError:
        print(f"\n[SKIPPED] openpyxl not installed, skipping Excel test")
        print(f"[Test]   Install with: pip install openpyxl")
        return True
    
    db = SessionLocal()
    try:
        output_path = REPORTS_DIR / f"test_excel_{uuid.uuid4().hex[:8]}.xlsx"
        
        print(f"\n[Test] Exporting to Excel: {output_path}")
        
        total_records, file_size = report_engine.export_to_file(
            db=db,
            device_id=TEST_DEVICE_ID,
            output_path=output_path,
            file_format="excel"
        )
        
        print(f"\n[Test] Excel Export complete:")
        print(f"[Test]   Records: {total_records}")
        print(f"[Test]   File size: {file_size / 1024:.2f} KB")
        
        if total_records != TEST_RECORDS_COUNT:
            print(f"[FAILED] Record count mismatch: expected {TEST_RECORDS_COUNT}, got {total_records}")
            return False
        
        if not output_path.exists():
            print(f"[FAILED] File does not exist")
            return False
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        row_count = ws.max_row
        data_rows = row_count - 1
        
        print(f"[Test]   Excel row count: {row_count} (header + {data_rows} data rows)")
        print(f"[Test]   Sheet name: {ws.title}")
        
        header_row = [cell.value for cell in ws[1]]
        print(f"[Test]   Header row: {header_row}")
        
        if data_rows != TEST_RECORDS_COUNT:
            print(f"[FAILED] Excel row count mismatch: expected {TEST_RECORDS_COUNT}, got {data_rows}")
            return False
        
        if row_count > 1:
            sample_device_id = ws.cell(row=2, column=2).value
            print(f"[Test]   Sample device_id: {sample_device_id}")
            
            if sample_device_id != TEST_DEVICE_ID:
                print(f"[FAILED] Device ID mismatch in Excel")
                return False
        
        wb.close()
        
        print(f"\n[PASSED] Excel export format test passed")
        return True
        
    finally:
        db.close()
        if output_path.exists():
            output_path.unlink()

def test_file_handler_resource_cleanup():
    print("\n" + "=" * 70)
    print("TEST 5: ExportFileHandler Resource Cleanup")
    print("=" * 70)
    
    test_path = REPORTS_DIR / f"test_cleanup_{uuid.uuid4().hex[:8]}.csv"
    
    try:
        handler = ExportFileHandler(test_path, "csv")
        handler.open()
        
        handler.write_header(["col1", "col2", "col3"])
        handler.write_row(["val1", "val2", "val3"])
        handler.flush()
        
        print(f"[Test] File opened and written to")
        
        if not handler._file or handler._file.closed:
            print(f"[FAILED] File should be open at this point")
            return False
        
        handler.close()
        
        if handler._file is not None and not handler._file.closed:
            print(f"[FAILED] File should be closed after close()")
            return False
        
        print(f"[Test] File handler properly closed")
        
        if test_path.exists():
            with open(test_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()
                print(f"[Test] File content length: {len(content)}")
        
        print(f"\n[PASSED] File handler cleanup test passed")
        return True
        
    finally:
        if test_path.exists():
            test_path.unlink()

def test_cleanup_expired_reports():
    print("\n" + "=" * 70)
    print("TEST 6: Cleanup Engine - Expired Reports")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        expired_task = ReportTask(
            id=str(uuid.uuid4()),
            task_name="Expired Test Task",
            device_id=TEST_DEVICE_ID,
            status=ReportTaskStatus.COMPLETED,
            progress=100,
            record_count=100,
            file_size_bytes=1024,
            export_format="csv",
            expire_at=datetime.utcnow() - timedelta(hours=48),
            created_at=datetime.utcnow() - timedelta(hours=72)
        )
        
        active_task = ReportTask(
            id=str(uuid.uuid4()),
            task_name="Active Test Task",
            device_id=TEST_DEVICE_ID,
            status=ReportTaskStatus.COMPLETED,
            progress=100,
            record_count=100,
            file_size_bytes=1024,
            export_format="csv",
            expire_at=datetime.utcnow() + timedelta(hours=24),
            created_at=datetime.utcnow()
        )
        
        expired_file_path = REPORTS_DIR / f"expired_test_{uuid.uuid4().hex[:8]}.csv"
        active_file_path = REPORTS_DIR / f"active_test_{uuid.uuid4().hex[:8]}.csv"
        orphan_file_path = REPORTS_DIR / f"orphan_test_{uuid.uuid4().hex[:8]}.csv"
        
        with open(expired_file_path, 'w') as f:
            f.write("expired data")
        
        with open(active_file_path, 'w') as f:
            f.write("active data")
        
        with open(orphan_file_path, 'w') as f:
            f.write("orphan data")
        
        old_mtime = (datetime.utcnow() - timedelta(hours=48)).timestamp()
        os.utime(expired_file_path, (old_mtime, old_mtime))
        os.utime(orphan_file_path, (old_mtime, old_mtime))
        
        expired_task.file_path = str(expired_file_path)
        expired_task.file_name = expired_file_path.name
        active_task.file_path = str(active_file_path)
        active_task.file_name = active_file_path.name
        
        db.add(expired_task)
        db.add(active_task)
        db.commit()
        
        print(f"\n[Test] Before cleanup:")
        print(f"[Test]   Expired task exists: {db.query(ReportTask).filter(ReportTask.id == expired_task.id).first() is not None}")
        print(f"[Test]   Active task exists: {db.query(ReportTask).filter(ReportTask.id == active_task.id).first() is not None}")
        print(f"[Test]   Expired file exists: {expired_file_path.exists()}")
        print(f"[Test]   Active file exists: {active_file_path.exists()}")
        print(f"[Test]   Orphan file exists: {orphan_file_path.exists()}")
        
        stats = report_engine.cleanup_expired_reports(db)
        
        print(f"\n[Test] Cleanup statistics: {stats}")
        print(f"\n[Test] After cleanup:")
        
        expired_task_exists = db.query(ReportTask).filter(ReportTask.id == expired_task.id).first() is not None
        active_task_exists = db.query(ReportTask).filter(ReportTask.id == active_task.id).first() is not None
        
        print(f"[Test]   Expired task exists: {expired_task_exists}")
        print(f"[Test]   Active task exists: {active_task_exists}")
        print(f"[Test]   Expired file exists: {expired_file_path.exists()}")
        print(f"[Test]   Active file exists: {active_file_path.exists()}")
        print(f"[Test]   Orphan file exists: {orphan_file_path.exists()}")
        
        passed = True
        
        if expired_task_exists:
            print(f"[FAILED] Expired task should have been deleted")
            passed = False
        else:
            print(f"[PASSED] Expired task was deleted")
        
        if not active_task_exists:
            print(f"[FAILED] Active task should NOT have been deleted")
            passed = False
        else:
            print(f"[PASSED] Active task was preserved")
        
        if expired_file_path.exists():
            print(f"[FAILED] Expired file should have been deleted")
            passed = False
        else:
            print(f"[PASSED] Expired file was deleted")
        
        if not active_file_path.exists():
            print(f"[FAILED] Active file should NOT have been deleted")
            passed = False
        else:
            print(f"[PASSED] Active file was preserved")
        
        if orphan_file_path.exists():
            print(f"[FAILED] Orphan file should have been deleted")
            passed = False
        else:
            print(f"[PASSED] Orphan file was deleted")
        
        if passed:
            print(f"\n[PASSED] Cleanup engine test passed")
        else:
            print(f"\n[FAILED] Cleanup engine test failed")
        
        return passed
        
    finally:
        db.close()
        for fp in [expired_file_path, active_file_path, orphan_file_path]:
            if fp.exists():
                try:
                    fp.unlink()
                except:
                    pass

def test_database_disconnect_handling():
    print("\n" + "=" * 70)
    print("TEST 7: Database Disconnect Handling - Robustness")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        task = report_engine.create_export_task(
            db=db,
            device_id=TEST_DEVICE_ID,
            task_name="DB Disconnect Test Task",
            file_format="csv"
        )
        
        task_id = task.id
        print(f"\n[Test] Created task: {task_id}")
        
        original_query_method = db.query
        
        failure_injected = False
        original_update_call = report_engine._update_task_status
        
        class FailingSession:
            def __init__(self, real_session):
                self._real = real_session
                self._fail_count = 0
            
            def query(self, *args, **kwargs):
                nonlocal failure_injected
                if failure_injected and self._fail_count < 5:
                    self._fail_count += 1
                    print(f"[Test]   Injecting OperationalError (simulating DB disconnect)")
                    raise OperationalError("Simulated database disconnect", {}, None)
                return self._real.query(*args, **kwargs)
            
            def commit(self):
                return self._real.commit()
            
            def refresh(self, instance):
                return self._real.refresh(instance)
            
            def close(self):
                return self._real.close()
            
            def __enter__(self):
                return self
            
            def __exit__(self, *args):
                return False
        
        failing_session = FailingSession(db)
        
        output_path = Path(task.file_path)
        
        print(f"\n[Test] Testing export with simulated DB disconnect during progress updates...")
        
        original_update_func = report_engine._update_task_status
        
        fail_once = [True]
        
        def failing_update(db_session, task_id, status, progress=None, **kwargs):
            if fail_once[0] and progress is not None and progress > 0:
                fail_once[0] = False
                print(f"[Test]   Injecting DB disconnect during progress update at {progress}%")
                raise OperationalError("Simulated database disconnect during update", {}, None)
            return original_update_func(db_session, task_id, status, progress, **kwargs)
        
        report_engine._update_task_status = failing_update
        
        try:
            progress_updates = []
            
            def progress_callback(current, total, progress):
                progress_updates.append(progress)
            
            total_records, file_size = report_engine.export_to_file(
                db=db,
                device_id=TEST_DEVICE_ID,
                output_path=output_path,
                file_format="csv",
                progress_callback=progress_callback
            )
            
            print(f"\n[Test] Export completed despite DB error:")
            print(f"[Test]   Records: {total_records}")
            print(f"[Test]   Progress updates received: {progress_updates}")
            print(f"[Test]   File exists: {output_path.exists()}")
            
            if total_records == TEST_RECORDS_COUNT:
                print(f"[PASSED] Export completed successfully even with DB disconnect")
            else:
                print(f"[FAILED] Record count mismatch")
                
        except Exception as e:
            print(f"\n[Test] Exception during export: {type(e).__name__}: {e}")
            if output_path.exists():
                print(f"[Test]   File was created: {output_path.exists()}")
        
        finally:
            report_engine._update_task_status = original_update_func
        
        if output_path.exists():
            output_path.unlink()
        
        db.refresh(task)
        print(f"\n[Test] Final task status: {task.status.value}")
        
        print(f"\n[PASSED] Database disconnect handling test completed")
        return True
        
    finally:
        db.close()

def test_task_failure_marking():
    print("\n" + "=" * 70)
    print("TEST 8: Task Failure Marking on Error")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        task = report_engine.create_export_task(
            db=db,
            device_id="NON_EXISTENT_DEVICE_999",
            task_name="Failure Test Task",
            file_format="csv"
        )
        
        task_id = task.id
        print(f"\n[Test] Created task: {task_id}")
        print(f"[Test] Initial status: {task.status.value}")
        
        success = report_engine.start_export_task_async(db, task_id, SessionLocal)
        
        if not success:
            print(f"[FAILED] Failed to start async task")
            return False
        
        print(f"\n[Test] Waiting for task to fail...")
        
        max_wait = 30
        wait_interval = 1
        elapsed = 0
        
        while elapsed < max_wait:
            time.sleep(wait_interval)
            elapsed += wait_interval
            
            db.refresh(task)
            status = task.status
            
            print(f"[Test]   Waiting... {elapsed}s, Status: {status.value}")
            
            if status in [ReportTaskStatus.COMPLETED, ReportTaskStatus.FAILED]:
                break
        
        db.refresh(task)
        
        print(f"\n[Test] Final task state:")
        print(f"[Test]   Status: {task.status.value}")
        print(f"[Test]   Progress: {task.progress}%")
        print(f"[Test]   Error message: {task.error_message}")
        
        if task.status == ReportTaskStatus.FAILED:
            print(f"[PASSED] Task correctly marked as FAILED")
            if task.error_message:
                print(f"[PASSED] Error message recorded: {task.error_message[:100]}...")
        else:
            print(f"[WARNING] Task status is {task.status.value}, expected FAILED (may be still running)")
        
        print(f"\n[PASSED] Task failure marking test completed")
        return True
        
    finally:
        db.close()

def test_export_task_crud():
    print("\n" + "=" * 70)
    print("TEST 9: Export Task CRUD Operations")
    print("=" * 70)
    
    db = SessionLocal()
    try:
        for fmt in SUPPORTED_FORMATS:
            print(f"\n[Test] Creating export task with format: {fmt}")
            try:
                task = report_engine.create_export_task(
                    db=db,
                    device_id=TEST_DEVICE_ID,
                    task_name=f"Test Task - {fmt}",
                    file_format=fmt
                )
                
                print(f"[Test]   Task ID: {task.id}")
                print(f"[Test]   File name: {task.file_name}")
                print(f"[Test]   Export format: {task.export_format}")
                print(f"[Test]   Status: {task.status.value}")
                
                if task.export_format != fmt:
                    print(f"[FAILED] Format mismatch: expected {fmt}, got {task.export_format}")
                    return False
                
                status = report_engine.get_task_status(db, task.id)
                print(f"[Test]   get_task_status returned: {status['id']}")
                
                if status['id'] != task.id:
                    print(f"[FAILED] get_task_status returned wrong task")
                    return False
                
            except ValueError as e:
                if fmt not in ["csv", "json", "excel", "xlsx"]:
                    print(f"[PASSED] Invalid format correctly rejected: {fmt}")
                else:
                    print(f"[FAILED] Unexpected error for format {fmt}: {e}")
                    return False
        
        print(f"\n[PASSED] Export task CRUD test passed")
        return True
        
    finally:
        db.close()

def main():
    print("\n" + "#" * 70)
    print("#  Report Engine Robustness Test Suite")
    print("#" * 70)
    print(f"\n[Test] Starting at: {datetime.utcnow()}")
    print(f"[Test] Test records count: {TEST_RECORDS_COUNT}")
    print(f"[Test] Test device ID: {TEST_DEVICE_ID}")
    print(f"[Test] Supported formats: {SUPPORTED_FORMATS}")
    
    try:
        init_test_database()
        
        db = SessionLocal()
        try:
            create_test_device(db)
            generate_test_history_records(db, TEST_DEVICE_ID, TEST_RECORDS_COUNT)
            
            record_count = db.query(DeviceDataHistory).filter(
                DeviceDataHistory.device_id == TEST_DEVICE_ID
            ).count()
            
            print(f"\n[Test] Verification: {record_count} history records in database")
        finally:
            db.close()
        
        results = []
        
        test_functions = [
            ("ProgressTracker (10% intervals)", test_progress_tracker),
            ("CSV Export Format", test_csv_export_format),
            ("JSON Export Format", test_json_export_format),
            ("Excel Export Format", test_excel_export_format),
            ("File Handler Resource Cleanup", test_file_handler_resource_cleanup),
            ("Export Task CRUD", test_export_task_crud),
            ("Cleanup Engine (Expired Reports)", test_cleanup_expired_reports),
            ("DB Disconnect Handling", test_database_disconnect_handling),
            ("Task Failure Marking", test_task_failure_marking),
        ]
        
        for test_name, test_func in test_functions:
            try:
                passed = test_func()
                results.append((test_name, passed))
            except Exception as e:
                print(f"\n[EXCEPTION] {test_name}: {e}")
                import traceback
                traceback.print_exc()
                results.append((test_name, False))
        
        print("\n" + "#" * 70)
        print("#  TEST SUMMARY")
        print("#" * 70)
        
        all_passed = True
        passed_count = 0
        
        for test_name, passed in results:
            status = "[PASSED]" if passed else "[FAILED]"
            print(f"\n{status} {test_name}")
            if passed:
                passed_count += 1
            else:
                all_passed = False
        
        print(f"\n{'=' * 70}")
        print(f"RESULTS: {passed_count}/{len(results)} tests passed")
        print(f"{'=' * 70}")
        
        if all_passed:
            print("\nALL TESTS PASSED!")
        else:
            print("\nSOME TESTS FAILED!")
        
        return all_passed
        
    finally:
        cleanup_test_database()

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

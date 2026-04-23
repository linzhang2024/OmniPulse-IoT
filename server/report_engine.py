import csv
import io
import os
import json
import uuid
import threading
import logging
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any, Generator, Tuple, Callable
from contextlib import contextmanager

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

from sqlalchemy.orm import Session
from sqlalchemy import and_, desc
from sqlalchemy.exc import SQLAlchemyError, OperationalError

logger = logging.getLogger("ReportEngine")

REPORTS_DIR = Path("./reports")
EXPORT_BATCH_SIZE = 1000
DEFAULT_REPORT_RETENTION_HOURS = 24
MAX_RECENT_TASKS = 100
PROGRESS_UPDATE_INTERVAL_PERCENT = 10

SUPPORTED_FORMATS = ["csv", "json", "excel", "xlsx"]

from server.models import (
    ReportTask, ReportTaskStatus, 
    ScheduledReportConfig, ScheduledReportType,
    DeviceDataHistory, Device, User
)

class ExportFileHandler:
    def __init__(self, file_path: Path, file_format: str):
        self.file_path = file_path
        self.file_format = file_format.lower()
        self._file = None
        self._writer = None
        self._workbook = None
        self._worksheet = None
        self._row_count = 0
        self._header_written = False
        self._json_records = []
        
    def __enter__(self):
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False
    
    def open(self):
        if self.file_format in ["csv"]:
            self._file = open(self.file_path, 'w', newline='', encoding='utf-8-sig')
            self._writer = csv.writer(self._file)
        elif self.file_format in ["json"]:
            pass
        elif self.file_format in ["excel", "xlsx"]:
            if not HAS_OPENPYXL:
                raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
            self._workbook = openpyxl.Workbook()
            self._worksheet = self._workbook.active
            self._worksheet.title = "Data Export"
    
    def write_header(self, headers: List[str]):
        if self._header_written:
            return
        
        if self.file_format == "csv":
            self._writer.writerow(headers)
        elif self.file_format == "json":
            self._json_headers = headers
        elif self.file_format in ["excel", "xlsx"]:
            if HAS_OPENPYXL:
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = self._worksheet.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    self._worksheet.column_dimensions[chr(64 + col_idx)].width = 18
        
        self._header_written = True
        self._row_count = 1
    
    def write_row(self, row: List[Any]):
        if self.file_format == "csv":
            self._writer.writerow(row)
        elif self.file_format == "json":
            record = {}
            for idx, key in enumerate(self._json_headers):
                record[key] = row[idx] if idx < len(row) else None
            self._json_records.append(record)
        elif self.file_format in ["excel", "xlsx"]:
            if HAS_OPENPYXL:
                self._row_count += 1
                for col_idx, value in enumerate(row, 1):
                    cell = self._worksheet.cell(row=self._row_count, column=col_idx, value=value)
        
        self._row_count += 0 if self.file_format in ["excel", "xlsx"] else 0
    
    def write_rows(self, rows: List[List[Any]]):
        for row in rows:
            self.write_row(row)
    
    def flush(self):
        if self._file and not self._file.closed:
            self._file.flush()
    
    def close(self):
        if self.file_format == "csv" and self._file:
            try:
                if not self._file.closed:
                    self._file.flush()
                    self._file.close()
            except Exception as e:
                logger.warning(f"[ExportFileHandler] Error closing CSV file: {e}")
        
        elif self.file_format == "json":
            try:
                with open(self.file_path, 'w', encoding='utf-8') as f:
                    json.dump(self._json_records, f, ensure_ascii=False, indent=2, default=str)
            except Exception as e:
                logger.warning(f"[ExportFileHandler] Error writing JSON file: {e}")
        
        elif self.file_format in ["excel", "xlsx"] and self._workbook:
            try:
                if HAS_OPENPYXL:
                    self._workbook.save(self.file_path)
            except Exception as e:
                logger.warning(f"[ExportFileHandler] Error saving Excel file: {e}")
        
        self._file = None
        self._writer = None
        self._workbook = None
        self._worksheet = None
        
        logger.debug(f"[ExportFileHandler] Closed file: {self.file_path}")

class ProgressTracker:
    def __init__(self, total: int, update_interval_percent: int = PROGRESS_UPDATE_INTERVAL_PERCENT):
        self.total = total
        self.update_interval_percent = update_interval_percent
        self.current = 0
        self.last_reported_percent = -1
        self._lock = threading.Lock()
    
    def update(self, current: int) -> Tuple[int, bool]:
        with self._lock:
            self.current = current
            if self.total == 0:
                return 0, False
            
            current_percent = int((current / self.total) * 100)
            
            should_update = False
            if self.last_reported_percent == -1:
                should_update = True
                self.last_reported_percent = 0
            elif current_percent - self.last_reported_percent >= self.update_interval_percent:
                should_update = True
                self.last_reported_percent = (current_percent // self.update_interval_percent) * self.update_interval_percent
            
            if current >= self.total and self.last_reported_percent < 100:
                should_update = True
                self.last_reported_percent = 100
            
            return current_percent, should_update
    
    def reset(self):
        with self._lock:
            self.current = 0
            self.last_reported_percent = -1

class ReportExportEngine:
    def __init__(self):
        self._export_lock = threading.Lock()
        self._active_tasks: Dict[str, threading.Thread] = {}
        self._task_callbacks: Dict[str, List] = {}
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        self._scheduled_configs: Dict[str, Dict] = {}
        self._cleanup_lock = threading.Lock()
        
    def ensure_reports_dir(self):
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    
    def _get_file_extension(self, file_format: str) -> str:
        fmt = file_format.lower()
        if fmt == "csv":
            return "csv"
        elif fmt == "json":
            return "json"
        elif fmt in ["excel", "xlsx"]:
            return "xlsx"
        return "csv"
    
    def _get_export_filename(
        self, 
        device_id: str, 
        file_format: str,
        timestamp: datetime = None
    ) -> str:
        if timestamp is None:
            timestamp = datetime.utcnow()
        
        safe_device_id = device_id.replace('/', '_').replace('\\', '_')
        ts_str = timestamp.strftime("%Y%m%d_%H%M%S")
        ext = self._get_file_extension(file_format)
        
        return f"report_{safe_device_id}_{ts_str}.{ext}"
    
    def _get_csv_headers(self, include_payload: bool = True) -> List[str]:
        headers = [
            "id",
            "device_id",
            "timestamp",
            "temperature",
            "humidity",
            "is_alert",
        ]
        if include_payload:
            headers.append("payload_json")
        return headers
    
    def _format_row(self, record: DeviceDataHistory, include_payload: bool = True) -> List[Any]:
        row = [
            record.id,
            record.device_id,
            record.timestamp.isoformat() if record.timestamp else "",
            record.temperature if record.temperature is not None else "",
            record.humidity if record.humidity is not None else "",
            "1" if record.is_alert else "0",
        ]
        if include_payload:
            row.append(json.dumps(record.payload, ensure_ascii=False) if record.payload else "")
        return row
    
    def stream_history_records(
        self,
        db: Session,
        device_id: str,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        batch_size: int = EXPORT_BATCH_SIZE
    ) -> Generator[Tuple[int, List[DeviceDataHistory]], None, None]:
        query = db.query(DeviceDataHistory).filter(
            DeviceDataHistory.device_id == device_id
        )
        
        if start_time:
            query = query.filter(DeviceDataHistory.timestamp >= start_time)
        if end_time:
            query = query.filter(DeviceDataHistory.timestamp <= end_time)
        
        total_count = query.count()
        logger.info(f"[ReportEngine] Starting export for device {device_id}: {total_count} records total")
        
        offset = 0
        while offset < total_count:
            try:
                batch = query.order_by(DeviceDataHistory.timestamp.asc()).offset(offset).limit(batch_size).all()
            except OperationalError as e:
                logger.error(f"[ReportEngine] Database connection error during streaming: {e}")
                raise
            except SQLAlchemyError as e:
                logger.error(f"[ReportEngine] SQL error during streaming: {e}")
                raise
            
            if not batch:
                break
            
            yield total_count, batch
            offset += len(batch)
    
    def export_to_file(
        self,
        db: Session,
        device_id: str,
        output_path: Path,
        file_format: str = "csv",
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        include_payload: bool = True,
        progress_callback: Optional[Callable[[int, int, int], None]] = None
    ) -> Tuple[int, int]:
        if file_format.lower() not in SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {file_format}. Supported formats: {SUPPORTED_FORMATS}")
        
        headers = self._get_csv_headers(include_payload)
        total_records = 0
        file_size = 0
        
        file_handler = None
        progress_tracker = None
        
        try:
            file_handler = ExportFileHandler(output_path, file_format)
            file_handler.open()
            
            first_batch = True
            total_count = 0
            for total_count, batch in self.stream_history_records(
                db, device_id, start_time, end_time
            ):
                if first_batch:
                    file_handler.write_header(headers)
                    progress_tracker = ProgressTracker(total_count)
                    first_batch = False
                
                formatted_rows = [self._format_row(record, include_payload) for record in batch]
                file_handler.write_rows(formatted_rows)
                
                total_records += len(batch)
                
                if progress_tracker and progress_callback:
                    progress_pct, should_update = progress_tracker.update(total_records)
                    if should_update:
                        try:
                            progress_callback(total_records, total_count, progress_pct)
                            file_handler.flush()
                            logger.info(f"[ReportEngine] Export progress: {total_records}/{total_count} ({progress_pct}%)")
                        except Exception as e:
                            logger.warning(f"[ReportEngine] Progress callback error: {e}")
                
                file_handler.flush()
            
            if total_records == 0:
                file_handler.write_header(headers)
            
        except Exception as e:
            logger.error(f"[ReportEngine] Export error: {e}", exc_info=True)
            raise
        finally:
            if file_handler:
                try:
                    file_handler.close()
                except Exception as e:
                    logger.warning(f"[ReportEngine] Error closing file handler: {e}")
        
        if output_path.exists():
            file_size = output_path.stat().st_size
        
        return total_records, file_size
    
    def export_to_stream(
        self,
        db: Session,
        device_id: str,
        file_format: str = "csv",
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        include_payload: bool = True
    ) -> Generator[str, None, None]:
        if file_format == "csv":
            yield from self._export_csv_stream(db, device_id, start_time, end_time, include_payload)
        elif file_format == "json":
            yield from self._export_json_stream(db, device_id, start_time, end_time, include_payload)
        else:
            raise ValueError(f"Streaming export not supported for format: {file_format}")
    
    def _export_csv_stream(
        self,
        db: Session,
        device_id: str,
        start_time: Optional[datetime],
        end_time: Optional[datetime],
        include_payload: bool
    ):
        headers = self._get_csv_headers(include_payload)
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate()
        
        for total_count, batch in self.stream_history_records(
            db, device_id, start_time, end_time
        ):
            for record in batch:
                row = self._format_row(record, include_payload)
                writer.writerow(row)
            yield output.getvalue()
            output.seek(0)
            output.truncate()
    
    def _export_json_stream(
        self,
        db: Session,
        device_id: str,
        start_time: Optional[datetime],
        end_time: Optional[datetime],
        include_payload: bool
    ):
        headers = self._get_csv_headers(include_payload)
        is_first = True
        
        yield "[\n"
        
        for total_count, batch in self.stream_history_records(
            db, device_id, start_time, end_time
        ):
            for record in batch:
                row = self._format_row(record, include_payload)
                record_dict = {}
                for idx, key in enumerate(headers):
                    record_dict[key] = row[idx] if idx < len(row) else None
                
                if not is_first:
                    yield ",\n"
                is_first = False
                
                yield json.dumps(record_dict, ensure_ascii=False, default=str)
        
        yield "\n]"
    
    def create_export_task(
        self,
        db: Session,
        device_id: str,
        task_name: Optional[str] = None,
        user_id: Optional[str] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        include_payload: bool = True,
        retention_hours: int = DEFAULT_REPORT_RETENTION_HOURS,
        file_format: str = "csv"
    ) -> ReportTask:
        task_id = str(uuid.uuid4())
        
        if file_format.lower() not in SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {file_format}. Supported: {SUPPORTED_FORMATS}")
        
        now = datetime.utcnow()
        file_name = self._get_export_filename(device_id, file_format, now)
        file_path = REPORTS_DIR / file_name
        
        expire_at = now + timedelta(hours=retention_hours)
        
        task = ReportTask(
            id=task_id,
            task_name=task_name or f"Export_{device_id}_{now.strftime('%Y%m%d_%H%M%S')}",
            device_id=device_id,
            user_id=user_id,
            status=ReportTaskStatus.PENDING,
            progress=0,
            file_name=file_name,
            file_path=str(file_path),
            expire_at=expire_at,
            filters={
                "start_time": start_time.isoformat() if start_time else None,
                "end_time": end_time.isoformat() if end_time else None,
                "include_payload": include_payload,
                "file_format": file_format
            },
            export_format=file_format.lower()
        )
        
        db.add(task)
        db.commit()
        db.refresh(task)
        
        logger.info(f"[ReportEngine] Created export task: {task_id} for device {device_id}, format: {file_format}")
        return task
    
    def _update_task_status(
        self,
        db: Session,
        task_id: str,
        status: ReportTaskStatus,
        progress: Optional[int] = None,
        record_count: Optional[int] = None,
        file_size_bytes: Optional[int] = None,
        error_message: Optional[str] = None,
        download_url: Optional[str] = None
    ) -> Optional[ReportTask]:
        max_retries = 3
        for attempt in range(max_retries):
            try:
                task = db.query(ReportTask).filter(ReportTask.id == task_id).first()
                if not task:
                    return None
                
                task.status = status
                if progress is not None:
                    task.progress = progress
                if record_count is not None:
                    task.record_count = record_count
                if file_size_bytes is not None:
                    task.file_size_bytes = file_size_bytes
                if error_message is not None:
                    task.error_message = error_message
                if download_url is not None:
                    task.download_url = download_url
                
                if status == ReportTaskStatus.PROCESSING and task.start_time is None:
                    task.start_time = datetime.utcnow()
                if status in [ReportTaskStatus.COMPLETED, ReportTaskStatus.FAILED]:
                    task.completed_at = datetime.utcnow()
                
                db.commit()
                db.refresh(task)
                return task
                
            except OperationalError as e:
                logger.warning(f"[ReportEngine] Database error updating task {task_id} (attempt {attempt+1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    import time
                    time.sleep(0.5 * (attempt + 1))
                else:
                    logger.error(f"[ReportEngine] Failed to update task {task_id} after {max_retries} attempts")
                    return None
            except Exception as e:
                logger.error(f"[ReportEngine] Error updating task {task_id}: {e}")
                return None
        
        return None
    
    def _execute_export_task(
        self,
        task_id: str,
        db_session_factory,
        device_id: str,
        file_path: Path,
        start_time: Optional[datetime],
        end_time: Optional[datetime],
        include_payload: bool,
        file_format: str
    ):
        db = db_session_factory()
        
        def progress_callback(current: int, total: int, progress_pct: int):
            try:
                self._update_task_status(db, task_id, ReportTaskStatus.PROCESSING, progress=progress_pct)
            except Exception as e:
                logger.warning(f"[ReportEngine] Progress update failed for task {task_id}: {e}")
        
        try:
            self._update_task_status(db, task_id, ReportTaskStatus.PROCESSING, progress=0)
            self.ensure_reports_dir()
            
            logger.info(f"[ReportEngine] Starting export task {task_id}: format={file_format}, path={file_path}")
            
            total_records, file_size = self.export_to_file(
                db=db,
                device_id=device_id,
                output_path=file_path,
                file_format=file_format,
                start_time=start_time,
                end_time=end_time,
                include_payload=include_payload,
                progress_callback=progress_callback
            )
            
            download_url = f"/api/reports/download/{task_id}"
            
            self._update_task_status(
                db, task_id, ReportTaskStatus.COMPLETED,
                progress=100,
                record_count=total_records,
                file_size_bytes=file_size,
                download_url=download_url
            )
            
            logger.info(f"[ReportEngine] Task {task_id} completed: {total_records} records, {file_size} bytes")
            
            self._trigger_callbacks(task_id, "completed", {
                "task_id": task_id,
                "record_count": total_records,
                "file_size": file_size,
                "format": file_format
            })
            
        except OperationalError as e:
            error_msg = f"Database connection error: {str(e)}"
            logger.error(f"[ReportEngine] Task {task_id} failed with DB error: {e}", exc_info=True)
            try:
                self._update_task_status(
                    db, task_id, ReportTaskStatus.FAILED,
                    error_message=error_msg
                )
                self._trigger_callbacks(task_id, "failed", {
                    "task_id": task_id,
                    "error": error_msg,
                    "error_type": "database"
                })
            except:
                pass
                
        except Exception as e:
            error_msg = str(e)
            logger.error(f"[ReportEngine] Task {task_id} failed: {e}", exc_info=True)
            try:
                self._update_task_status(
                    db, task_id, ReportTaskStatus.FAILED,
                    error_message=error_msg
                )
                self._trigger_callbacks(task_id, "failed", {
                    "task_id": task_id,
                    "error": error_msg
                })
            except:
                pass
                
        finally:
            try:
                db.close()
            except Exception as e:
                logger.warning(f"[ReportEngine] Error closing DB session: {e}")
            
            with self._export_lock:
                if task_id in self._active_tasks:
                    del self._active_tasks[task_id]
    
    def start_export_task_async(
        self,
        db: Session,
        task_id: str,
        db_session_factory
    ) -> bool:
        task = db.query(ReportTask).filter(ReportTask.id == task_id).first()
        if not task:
            return False
        
        if task.status != ReportTaskStatus.PENDING:
            logger.warning(f"[ReportEngine] Task {task_id} is not in PENDING state: {task.status}")
            return False
        
        filters = task.filters or {}
        start_time = None
        if filters.get("start_time"):
            try:
                start_time = datetime.fromisoformat(filters["start_time"])
            except:
                pass
        
        end_time = None
        if filters.get("end_time"):
            try:
                end_time = datetime.fromisoformat(filters["end_time"])
            except:
                pass
        
        include_payload = filters.get("include_payload", True)
        file_format = filters.get("file_format", "csv") or task.export_format or "csv"
        
        file_path = Path(task.file_path) if task.file_path else None
        
        if not file_path:
            self._update_task_status(
                db, task_id, ReportTaskStatus.FAILED,
                error_message="Invalid file path"
            )
            return False
        
        self.ensure_reports_dir()
        
        thread = threading.Thread(
            target=self._execute_export_task,
            args=(task_id, db_session_factory, task.device_id, file_path, start_time, end_time, include_payload, file_format),
            daemon=True
        )
        
        with self._export_lock:
            self._active_tasks[task_id] = thread
        
        thread.start()
        logger.info(f"[ReportEngine] Started async export task: {task_id}")
        return True
    
    def get_task_status(self, db: Session, task_id: str) -> Optional[Dict]:
        task = db.query(ReportTask).filter(ReportTask.id == task_id).first()
        if not task:
            return None
        
        return {
            "id": task.id,
            "task_name": task.task_name,
            "device_id": task.device_id,
            "status": task.status.value,
            "progress": task.progress,
            "record_count": task.record_count,
            "file_size_bytes": task.file_size_bytes,
            "file_name": task.file_name,
            "download_url": task.download_url,
            "error_message": task.error_message,
            "export_format": task.export_format,
            "start_time": task.start_time.isoformat() if task.start_time else None,
            "completed_at": task.completed_at.isoformat() if task.completed_at else None,
            "expire_at": task.expire_at.isoformat() if task.expire_at else None,
            "created_at": task.created_at.isoformat() if task.created_at else None
        }
    
    def cleanup_expired_reports(self, db: Session) -> Dict[str, int]:
        with self._cleanup_lock:
            now = datetime.utcnow()
            stats = {
                "deleted_files": 0,
                "deleted_tasks": 0,
                "orphaned_files": 0,
                "errors": 0
            }
            
            logger.info(f"[ReportEngine] Starting cleanup at {now}")
            
            try:
                expired_tasks = db.query(ReportTask).filter(
                    ReportTask.expire_at <= now,
                    ReportTask.status.in_([ReportTaskStatus.COMPLETED, ReportTaskStatus.FAILED])
                ).all()
                
                for task in expired_tasks:
                    try:
                        if task.file_path and os.path.exists(task.file_path):
                            try:
                                os.unlink(task.file_path)
                                stats["deleted_files"] += 1
                                logger.info(f"[ReportEngine] Removed expired report file: {task.file_path}")
                            except Exception as e:
                                logger.warning(f"[ReportEngine] Failed to delete file {task.file_path}: {e}")
                                stats["errors"] += 1
                        
                        db.delete(task)
                        stats["deleted_tasks"] += 1
                        logger.info(f"[ReportEngine] Deleted expired task: {task.id}")
                        
                    except Exception as e:
                        logger.warning(f"[ReportEngine] Failed to cleanup task {task.id}: {e}")
                        stats["errors"] += 1
                
                if stats["deleted_tasks"] > 0:
                    db.commit()
                    logger.info(f"[ReportEngine] Committed {stats['deleted_tasks']} task deletions")
                
            except Exception as e:
                logger.error(f"[ReportEngine] Error during task cleanup: {e}")
                stats["errors"] += 1
            
            try:
                self.ensure_reports_dir()
                
                for file_path in REPORTS_DIR.iterdir():
                    if not file_path.is_file():
                        continue
                    
                    try:
                        file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                        file_age_hours = (now - file_mtime).total_seconds() / 3600
                        
                        if file_age_hours > DEFAULT_REPORT_RETENTION_HOURS:
                            task_exists = db.query(ReportTask).filter(
                                ReportTask.file_path == str(file_path)
                            ).first()
                            
                            if not task_exists:
                                try:
                                    file_path.unlink()
                                    stats["orphaned_files"] += 1
                                    logger.info(f"[ReportEngine] Removed orphaned file: {file_path} (age: {file_age_hours:.1f}h)")
                                except Exception as e:
                                    logger.warning(f"[ReportEngine] Failed to delete orphaned file {file_path}: {e}")
                                    stats["errors"] += 1
                                    
                    except Exception as e:
                        logger.warning(f"[ReportEngine] Error checking file {file_path}: {e}")
                        stats["errors"] += 1
                        
            except Exception as e:
                logger.error(f"[ReportEngine] Error during orphaned file cleanup: {e}")
                stats["errors"] += 1
            
            logger.info(f"[ReportEngine] Cleanup completed: {stats}")
            return stats
    
    def register_callback(self, task_id: str, callback: callable, event: str = "completed"):
        with self._export_lock:
            if task_id not in self._task_callbacks:
                self._task_callbacks[task_id] = []
            self._task_callbacks[task_id].append({
                "callback": callback,
                "event": event
            })
    
    def _trigger_callbacks(self, task_id: str, event: str, data: Dict):
        with self._export_lock:
            callbacks = self._task_callbacks.get(task_id, [])
        
        for cb_info in callbacks:
            if cb_info["event"] == event or cb_info["event"] == "all":
                try:
                    cb_info["callback"](event, data)
                except Exception as e:
                    logger.warning(f"[ReportEngine] Callback error for task {task_id}: {e}")
        
        with self._export_lock:
            if task_id in self._task_callbacks:
                del self._task_callbacks[task_id]
    
    def get_active_tasks_count(self) -> int:
        with self._export_lock:
            return len(self._active_tasks)
    
    def cancel_task(self, db: Session, task_id: str) -> bool:
        task = db.query(ReportTask).filter(ReportTask.id == task_id).first()
        if not task:
            return False
        
        if task.status in [ReportTaskStatus.COMPLETED, ReportTaskStatus.FAILED]:
            return False
        
        with self._export_lock:
            if task_id in self._active_tasks:
                thread = self._active_tasks[task_id]
                if thread.is_alive():
                    logger.warning(f"[ReportEngine] Cannot forcefully cancel running thread for task {task_id}")
        
        self._update_task_status(
            db, task_id, ReportTaskStatus.FAILED,
            error_message="Task cancelled by user"
        )
        
        logger.info(f"[ReportEngine] Task {task_id} marked as cancelled")
        return True

class ScheduledReportManager:
    def __init__(self, report_engine: ReportExportEngine, scheduler=None):
        self.engine = report_engine
        self.scheduler = scheduler
        self._configs_cache: Dict[str, ScheduledReportConfig] = {}
        self._job_ids: Dict[str, str] = {}
        
    def load_active_configs(self, db: Session):
        active_configs = db.query(ScheduledReportConfig).filter(
            ScheduledReportConfig.is_active == True
        ).all()
        
        for config in active_configs:
            self._configs_cache[config.id] = config
        
        logger.info(f"[ScheduledReportManager] Loaded {len(self._configs_cache)} active scheduled report configs")
    
    def _create_default_daily_config(self, db: Session) -> ScheduledReportConfig:
        config = ScheduledReportConfig(
            id=str(uuid.uuid4()),
            name="每日运行简报",
            description="每日 8 点自动生成昨日设备运行简报",
            report_type=ScheduledReportType.DAILY_BRIEFING,
            cron_expression="0 8 * * *",
            output_directory="./reports/daily",
            retention_days=7,
            file_name_template="daily_briefing_{date}.csv",
            is_active=True
        )
        db.add(config)
        db.commit()
        db.refresh(config)
        return config
    
    def ensure_daily_briefing_config(self, db: Session) -> ScheduledReportConfig:
        existing = db.query(ScheduledReportConfig).filter(
            ScheduledReportConfig.report_type == ScheduledReportType.DAILY_BRIEFING
        ).first()
        
        if existing:
            return existing
        
        return self._create_default_daily_config(db)
    
    def generate_daily_briefing_filename(self, date: datetime, template: str = None, format: str = "csv") -> str:
        ext = self.engine._get_file_extension(format)
        if template:
            base_name = template.format(date=date.strftime("%Y%m%d"))
            if not base_name.endswith(f".{ext}"):
                base_name = f"{base_name}.{ext}"
            return base_name
        return f"daily_briefing_{date.strftime('%Y%m%d')}.{ext}"
    
    def _execute_scheduled_report(self, config_id: str, db_session_factory):
        db = db_session_factory()
        try:
            config = db.query(ScheduledReportConfig).filter(
                ScheduledReportConfig.id == config_id
            ).first()
            
            if not config or not config.is_active:
                logger.warning(f"[ScheduledReport] Config {config_id} not found or inactive")
                return
            
            config.last_run_at = datetime.utcnow()
            config.last_run_status = "running"
            db.commit()
            
            logger.info(f"[ScheduledReport] Executing {config.report_type.value}: {config.name}")
            
            if config.report_type == ScheduledReportType.DAILY_BRIEFING:
                self._generate_daily_briefing(db, config, db_session_factory)
            else:
                logger.warning(f"[ScheduledReport] Unsupported report type: {config.report_type}")
            
            config.last_run_status = "success"
            config.last_run_error = None
            db.commit()
            logger.info(f"[ScheduledReport] Completed {config.name}")
            
        except Exception as e:
            logger.error(f"[ScheduledReport] Failed to execute report {config_id}: {e}", exc_info=True)
            try:
                config = db.query(ScheduledReportConfig).filter(
                    ScheduledReportConfig.id == config_id
                ).first()
                if config:
                    config.last_run_status = "failed"
                    config.last_run_error = str(e)
                    db.commit()
            except:
                pass
        finally:
            try:
                db.close()
            except:
                pass
    
    def _generate_daily_briefing(
        self, 
        db: Session, 
        config: ScheduledReportConfig,
        db_session_factory
    ):
        now = datetime.utcnow()
        yesterday = now - timedelta(days=1)
        start_of_yesterday = datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0)
        end_of_yesterday = datetime(yesterday.year, yesterday.month, yesterday.day, 23, 59, 59)
        
        output_dir = Path(config.output_directory or "./reports/daily")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        export_format = config.export_format or "csv"
        
        devices = db.query(Device).all()
        
        for device in devices:
            history_count = db.query(DeviceDataHistory).filter(
                DeviceDataHistory.device_id == device.device_id,
                DeviceDataHistory.timestamp >= start_of_yesterday,
                DeviceDataHistory.timestamp <= end_of_yesterday
            ).count()
            
            if history_count == 0:
                continue
            
            file_name = self.generate_daily_briefing_filename(yesterday, config.file_name_template, export_format)
            file_path = output_dir / file_name
            
            task = self.engine.create_export_task(
                db,
                device_id=device.device_id,
                task_name=f"每日简报_{device.device_id}_{yesterday.strftime('%Y-%m-%d')}",
                start_time=start_of_yesterday,
                end_time=end_of_yesterday,
                include_payload=True,
                retention_hours=config.retention_days * 24,
                file_format=export_format
            )
            
            task.file_path = str(file_path)
            task.file_name = file_name
            db.commit()
            
            self.engine.start_export_task_async(db, task.id, db_session_factory)
            
            logger.info(f"[DailyBriefing] Created export task for device {device.device_id}: {task.id}")
    
    def schedule_all_active(self, db: Session):
        if not self.scheduler:
            logger.warning("[ScheduledReportManager] No scheduler available, skipping scheduling")
            return
        
        active_configs = db.query(ScheduledReportConfig).filter(
            ScheduledReportConfig.is_active == True
        ).all()
        
        for config in active_configs:
            self.schedule_config(config, db)
    
    def schedule_config(self, config: ScheduledReportConfig, db: Session):
        if not self.scheduler:
            return
        
        if config.id in self._job_ids:
            try:
                self.scheduler.remove_job(self._job_ids[config.id])
            except:
                pass
        
        cron_expr = config.cron_expression or "0 8 * * *"
        
        try:
            from apscheduler.triggers.cron import CronTrigger
            
            cron_parts = cron_expr.split()
            if len(cron_parts) == 5:
                trigger = CronTrigger(
                    minute=cron_parts[0],
                    hour=cron_parts[1],
                    day=cron_parts[2],
                    month=cron_parts[3],
                    day_of_week=cron_parts[4]
                )
            else:
                trigger = CronTrigger(hour=8, minute=0)
            
            job = self.scheduler.add_job(
                self._execute_scheduled_report,
                trigger=trigger,
                args=[config.id, lambda: db.session_factory() if hasattr(db, 'session_factory') else None],
                id=f"scheduled_report_{config.id}",
                name=f"Scheduled Report: {config.name}",
                replace_existing=True
            )
            
            self._job_ids[config.id] = job.id
            logger.info(f"[ScheduledReportManager] Scheduled {config.name} with cron: {cron_expr}")
            
        except Exception as e:
            logger.error(f"[ScheduledReportManager] Failed to schedule config {config.id}: {e}")

report_engine = ReportExportEngine()
